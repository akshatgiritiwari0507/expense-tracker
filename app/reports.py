from datetime import datetime, date
from flask import current_app
from sqlalchemy import func, extract, text
from app.models import Expense, Income, Category
import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import plotly.graph_objects as go
import plotly.express as px
from plotly.utils import PlotlyJSONEncoder
import json

class ReportGenerator:
    """Class for generating financial reports and analytics"""
    
    def __init__(self, user_id):
        self.user_id = user_id
    
    def get_monthly_expenses(self, year=None):
        """Get monthly expense summary"""
        if year is None:
            year = datetime.now().year
        
        from app import db
        monthly_data = db.session.query(
            extract('month', Expense.date).label('month'),
            func.sum(Expense.amount).label('total')
        ).filter(
            Expense.user_id == self.user_id,
            extract('year', Expense.date) == year
        ).group_by(
            extract('month', Expense.date)
        ).order_by('month').all()
        
        return [{'month': month, 'total': total} for month, total in monthly_data]
    
    def get_category_breakdown(self, start_date=None, end_date=None):
        """Get expense breakdown by category"""
        from app import db
        query = db.session.query(
            Expense.category,
            func.sum(Expense.amount).label('total'),
            func.count(Expense.id).label('count')
        ).filter(Expense.user_id == self.user_id)
        
        if start_date:
            query = query.filter(Expense.date >= start_date)
        if end_date:
            query = query.filter(Expense.date <= end_date)
        
        category_data = query.group_by(Expense.category).all()
        
        return [{'category': cat, 'total': total, 'count': count} for cat, total, count in category_data]
    
    def get_spending_trends(self, months=6):
        """Get spending trends over specified months"""
        end_date = datetime.now().date()
        start_date = end_date.replace(month=end_date.month - months) if end_date.month > months else end_date.replace(year=end_date.year - 1, month=12 + end_date.month - months)
        
        from app import db
        trend_data = db.session.query(
            func.strftime('%Y-%m', Expense.date).label('period'),
            func.sum(Expense.amount).label('total'),
            func.count(Expense.id).label('count')
        ).filter(
            Expense.user_id == self.user_id,
            Expense.date >= start_date,
            Expense.date <= end_date
        ).group_by('period').order_by('period').all()
        
        return [{'period': period, 'total': total, 'count': count} for period, total, count in trend_data]
    
    def generate_pdf_report(self, start_date=None, end_date=None, report_type='monthly'):
        """Generate PDF financial report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph(f"Financial Report - {report_type.title()}", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Date range
        date_range = f"Period: {start_date or 'All time'} to {end_date or 'Present'}"
        story.append(Paragraph(date_range, styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Category breakdown table
        category_data = self.get_category_breakdown(start_date, end_date)
        if category_data:
            table_data = [['Category', 'Total Amount', 'Transaction Count']]
            for cat in category_data:
                table_data.append([cat['category'], f"${cat['total']:.2f}", str(cat['count'])])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_excel_report(self, start_date=None, end_date=None):
        """Generate Excel financial report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"
        
        # Header
        ws['A1'] = "Financial Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Period: {start_date or 'All time'} to {end_date or 'Present'}"
        
        # Category breakdown
        category_data = self.get_category_breakdown(start_date, end_date)
        if category_data:
            ws['A4'] = "Category Breakdown"
            ws['A4'].font = Font(bold=True)
            
            headers = ['Category', 'Total Amount', 'Transaction Count']
            ws.append(headers)
            
            for cat in category_data:
                ws.append([cat['category'], cat['total'], cat['count']])
        
        # Style the header row
        for col in range(1, 4):
            cell = ws.cell(row=5, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_spending_chart(self, chart_type='pie'):
        """Generate spending visualization"""
        category_data = self.get_category_breakdown()
        
        if not category_data:
            return None
        
        if chart_type == 'pie':
            fig = go.Figure(data=[go.Pie(
                labels=[cat['category'] for cat in category_data],
                values=[cat['total'] for cat in category_data],
                title="Expense Distribution by Category"
            )])
        elif chart_type == 'bar':
            fig = go.Figure(data=[
                go.Bar(
                    x=[cat['category'] for cat in category_data],
                    y=[cat['total'] for cat in category_data],
                    text=[f"${cat['total']:.2f}" for cat in category_data],
                    textposition='auto'
                )
            ])
            fig.update_layout(
                title="Expenses by Category",
                xaxis_title="Category",
                yaxis_title="Amount ($)"
            )
        
        fig.update_layout(
            font=dict(size=14),
            height=500
        )
        
        return json.dumps(fig, cls=PlotlyJSONEncoder)
    
    def generate_trend_chart(self):
        """Generate spending trend chart"""
        trend_data = self.get_spending_trends()
        
        if not trend_data:
            return None
        
        fig = go.Figure()
        
        fig.add_trace(go.Scatter(
            x=[t['period'] for t in trend_data],
            y=[t['total'] for t in trend_data],
            mode='lines+markers',
            name='Monthly Spending',
            line=dict(color='blue', width=3),
            marker=dict(size=8)
        ))
        
        fig.update_layout(
            title="Spending Trends (Last 6 Months)",
            xaxis_title="Period",
            yaxis_title="Amount ($)",
            font=dict(size=14),
            height=500
        )
        
        return json.dumps(fig, cls=PlotlyJSONEncoder)
